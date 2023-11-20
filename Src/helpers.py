import pandas as pd
import openpyxl as op
import sys
import os

# Hash function to convert Letters to Numbers (Only one letter)
def letter_to_number(letter):
    if len(letter) == 1:
        return ord(letter.lower()) - 97 


# Function to help with paths
def resource_path(relative_path):
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)


# One long function to rule them all (and update prices from excel A to B using given codes)
def update_prices(path_proveedor, 
                  path_local, 
                  ubicacion_codigo_proveedor,
                  ubicacion_precio_proveedor, 
                  ubicacion_codigo_local, 
                  ubicacion_precio_local):

    data_empresa = pd.read_excel(path_proveedor)
    data_ferreteria = pd.read_excel(path_local)

    columnas = list(data_ferreteria.columns.values)

    # Sacar nombre de categoría en columna 1
    data_ferreteria.columns.values[1]

    # Pido las Celdas de la info que necesito
    valor_codigo_empresa = ubicacion_codigo_proveedor
    valor_precio_empresa = ubicacion_precio_proveedor
    valor_codigo_ferreteria = ubicacion_codigo_local
    valor_precio_ferreteria = ubicacion_precio_local

    # Me quedo con las columnas que me interesan del excel de la empresa
    data_empresa_importante = data_empresa[[data_empresa.columns.values[letter_to_number(valor_codigo_empresa)], data_empresa.columns.values[letter_to_number(valor_precio_empresa)]]]
    data_ferreteria_importante = data_ferreteria[ data_ferreteria.columns.values[letter_to_number(valor_codigo_ferreteria)]]


    merge = pd.merge(data_ferreteria_importante, 
                    data_empresa_importante, 
                    left_on = data_ferreteria.columns.values[letter_to_number(valor_codigo_ferreteria)],
                    right_on = data_empresa.columns.values[letter_to_number(valor_codigo_empresa)])


    merge.to_excel(resource_path("prueba.xlsx"))
    resultados = pd.read_excel(resource_path("prueba.xlsx"))

    # Crear un diccionario que contenga el codigo como key y el precio actualizado como valor
    dic_resultados = dict()
    print(resultados)
    for i in range(len(resultados)):
        dic_resultados[resultados.iloc[i,1]] = resultados.iloc[i,3]

    # Cambiar los datos en df de la ferretería
    #data_ferreteria[data_ferreteria.columns.values[letter_to_number(valor_precio_ferreteria)]] = data_ferreteria[data_ferreteria.columns.values[letter_to_number(valor_codigo_ferreteria)]].map(dic_resultados)

    #data_ferreteria.to_excel("FerreteriaNorte\precios_finales.xlsx")

    
    # Write new data into existing file
    wb = op.load_workbook(path_local)
    ws = wb.active
    total_rows = ws.max_row

    for key in dic_resultados.keys():
        for j in range(1, total_rows):
            celda_codigo = valor_codigo_ferreteria.upper() + str(j)
            if ws[celda_codigo].value == key:
                celda_precio = valor_precio_ferreteria.upper() + str(j)
                ws[celda_precio].value = dic_resultados[key]

    wb.save(path_local)


# Function to search a given code in excel file
def search_item(data, column, code):
    for i in range(len(data)):
        if (data.loc[i][column] == code):
            return True
    return False


# Function to search for products availability by code from a given excel file from ferretería NORTE
def filter_products(consumer_data, stock_data, consumer_column, stock_column):
    products_availability = {"available":[], "not available": []}

    for i in range(len(consumer_data)):
        if (search_item(stock_data, stock_column, consumer_data.loc[i][consumer_column])):
            products_availability["available"].append(consumer_data.loc[i][consumer_column])
        else:
            products_availability["not available"].append(consumer_data.loc[i][consumer_column])
    return products_availability

# Function to search a given code in CSV file
def search_item_csv(data, column, code):
    for i in range(len(data)):
        if (data.iloc[i, column] == code):
            return True
    return False


# Function to search for products availability by code from a given excel CSV from ferretería NORTE
def filter_products_csv(consumer_data, stock_data, consumer_column, stock_column):
    products_availability = {"available":[], "not available": []}

    for i in range(len(consumer_data)):
        if (search_item(stock_data, stock_column, consumer_data.iloc[i, consumer_column])):
            products_availability["available"].append(consumer_data.iloc[i, consumer_column])
        else:
            products_availability["not available"].append(consumer_data.iloc[i, consumer_column])
    return products_availability