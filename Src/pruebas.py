import openpyxl as op

# Write new data into existing file
wb = op.load_workbook("FerreteriaNorte\Fer_actualizar.xlsx")
ws = wb.active

total_rows = ws.max_row

#for i in range(1, total_rows):
#   print(ws['C' + str(i)].value)

ws['I2'].value = 980

wb.save("FerreteriaNorte\Fer_actualizar.xlsx")