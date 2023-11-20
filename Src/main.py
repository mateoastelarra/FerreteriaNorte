import pandas as pd
import openpyxl as op
from helpers import filter_products_csv, letter_to_number

data_empresa = pd.read_excel("FerreteriaNorte/Precios_nuevos.xlsx")
data_ferreteria = pd.read_excel("FerreteriaNorte\Fer_actualizar.xlsx")


columnas = list(data_ferreteria.columns.values)

# Sacar nombre de categoría en columna 1
data_ferreteria.columns.values[1]

# Pido las Celdas de la info que necesito
valor_codigo_empresa = input("Por Favor Ingrese la columna donde está el codigo en el excel de la empresa:")
valor_precio_empresa = input("Por Favor Ingrese la columna donde está el precio en el excel de la empresa:")
valor_codigo_ferreteria = input("Por Favor Ingrese la columna donde está el código en el excel de la ferretería:")
valor_precio_ferreteria = input("Por Favor Ingrese la columna donde está el precio a actualizar en el excel de la ferretería:")
print(data_ferreteria)

# Me quedo con las columnas que me interesan del excel de la empresa
data_empresa_importante = data_empresa[[data_empresa.columns.values[letter_to_number(valor_codigo_empresa)], data_empresa.columns.values[letter_to_number(valor_precio_empresa)]]]
data_ferreteria_importante = data_ferreteria[ data_ferreteria.columns.values[letter_to_number(valor_codigo_ferreteria)]]



merge = pd.merge(data_ferreteria_importante, 
                 data_empresa_importante, 
                 left_on = data_ferreteria.columns.values[letter_to_number(valor_codigo_ferreteria)],
                 right_on = data_empresa.columns.values[letter_to_number(valor_codigo_empresa)])


merge.to_excel("FerreteriaNorte\prueba.xlsx")
resultados = pd.read_excel("FerreteriaNorte\prueba.xlsx")


# Crear un diccionario que contenga el codigo como key y el precio actualizado como valor
dic_resultados = dict()

for i in range(len(resultados)):
    dic_resultados[resultados.iloc[i,1]] = resultados.iloc[i,3]

# Cambiar los datos en df de la ferretería
data_ferreteria[data_ferreteria.columns.values[letter_to_number(valor_precio_ferreteria)]] = data_ferreteria[data_ferreteria.columns.values[letter_to_number(valor_codigo_ferreteria)]].map(dic_resultados)

data_ferreteria.to_excel("FerreteriaNorte\precios_finales.xlsx")


# Write new data into existing file
wb = op.load_workbook("FerreteriaNorte\Fer_actualizar.xlsx")
ws = wb.active
total_rows = ws.max_row

for key in dic_resultados.keys():
   for j in range(1, total_rows):
       celda_codigo = valor_codigo_ferreteria.upper() + str(j)
       if ws[celda_codigo].value == key:
           celda_precio = valor_precio_ferreteria.upper() + str(j)
           ws[celda_precio].value = dic_resultados[key]

wb.save("FerreteriaNorte\Fer_actualizar.xlsx")